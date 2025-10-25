"""
Unit tests for excel_generate_v2.py
"""

import json
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import pytest
import openpyxl

# Add parent directory to path to import the module
sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_generate_v2 import (
    load_config,
    parse_date,
    set_cell_style,
    generate_excel,
    create_buildings_sheet,
    create_units_sheet,
    create_tenants_sheet,
    create_rents_sheet,
    create_expenses_sheet,
)


class TestLoadConfig:
    """Tests for load_config function"""

    def test_load_valid_config(self, tmp_path):
        """Test loading a valid configuration file"""
        config_data = {
            "output_filename": "test.xlsx",
            "buildings": [],
            "units": [],
            "tenants": [],
            "rents_paid": [],
            "expenses": []
        }
        config_file = tmp_path / "config.json"
        config_file.write_text(json.dumps(config_data, ensure_ascii=False))

        result = load_config(config_file)
        assert result == config_data

    def test_load_nonexistent_config(self, tmp_path):
        """Test loading a non-existent configuration file"""
        config_file = tmp_path / "nonexistent.json"

        with pytest.raises(FileNotFoundError) as exc_info:
            load_config(config_file)
        assert "Configuration file not found" in str(exc_info.value)

    def test_load_invalid_json(self, tmp_path):
        """Test loading an invalid JSON file"""
        config_file = tmp_path / "invalid.json"
        config_file.write_text("{ invalid json }")

        with pytest.raises(json.JSONDecodeError):
            load_config(config_file)


class TestParseDate:
    """Tests for parse_date function"""

    def test_parse_valid_date(self):
        """Test parsing a valid date string"""
        result = parse_date("2024-01-15")
        expected = datetime(2024, 1, 15).date()
        assert result == expected

    def test_parse_none_date(self):
        """Test parsing None"""
        result = parse_date(None)
        assert result is None

    def test_parse_empty_string(self):
        """Test parsing empty string"""
        result = parse_date("")
        assert result is None

    def test_parse_invalid_date(self, capsys):
        """Test parsing invalid date format"""
        result = parse_date("invalid-date")
        assert result is None
        captured = capsys.readouterr()
        assert "Warning: Invalid date format" in captured.err


class TestSetCellStyle:
    """Tests for set_cell_style function"""

    def test_set_cell_style_defaults(self):
        """Test setting cell style with default values"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        cell = sheet['A1']

        set_cell_style(cell)

        assert cell.font.size == 12
        assert cell.font.bold is False
        assert cell.alignment.horizontal == "center"
        assert cell.border.left.style == "thin"

    def test_set_cell_style_bold(self):
        """Test setting cell style with bold font"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        cell = sheet['A1']

        set_cell_style(cell, bold=True)

        assert cell.font.bold is True

    def test_set_cell_style_with_color(self):
        """Test setting cell style with background color"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        cell = sheet['A1']

        set_cell_style(cell, bg_color="FF0000")

        # openpyxl uses ARGB format (includes alpha channel)
        assert cell.fill.start_color.rgb == "00FF0000"


class TestGenerateExcel:
    """Tests for generate_excel function"""

    def test_generate_excel_basic(self, tmp_path):
        """Test generating a basic Excel file"""
        config = {
            "buildings": [{"name": "Test Building", "units": 5, "notes": ""}],
            "units": [],
            "tenants": [],
            "rents_paid": [],
            "expenses": []
        }
        output_file = tmp_path / "test_output.xlsx"

        generate_excel(config, output_file)

        assert output_file.exists()

        # Verify the Excel file can be opened
        wb = openpyxl.load_workbook(output_file)
        assert 'العمارات' in wb.sheetnames
        assert 'الوحدات' in wb.sheetnames
        assert 'المستأجرين' in wb.sheetnames
        assert 'الإيجارات' in wb.sheetnames
        assert 'المصروفات' in wb.sheetnames

    def test_generate_excel_with_data(self, tmp_path):
        """Test generating Excel file with actual data"""
        config = {
            "buildings": [{"name": "عمارة أ", "units": 2, "notes": ""}],
            "units": [
                {
                    "unit_no": "101",
                    "building": "عمارة أ",
                    "type": "شقة",
                    "rent": 5000,
                    "status": "مُؤجّرة",
                    "notes": ""
                }
            ],
            "tenants": [
                {
                    "unit_no": "101",
                    "name": "محمد",
                    "id": "123456",
                    "mobile": "0500000000",
                    "start_date": "2024-01-01",
                    "end_date": "2024-12-31",
                    "rent": 5000,
                    "email": "test@example.com",
                    "notes": ""
                }
            ],
            "rents_paid": [
                {
                    "unit_no": "101",
                    "month": "يناير",
                    "year": 2024,
                    "amount": 5000,
                    "date": "2024-01-05",
                    "method": "تحويل",
                    "status": "مدفوع",
                    "notes": ""
                }
            ],
            "expenses": [
                {
                    "building": "عمارة أ",
                    "date": "2024-01-01",
                    "type": "صيانة",
                    "amount": 500,
                    "category": "صيانة",
                    "notes": ""
                }
            ]
        }
        output_file = tmp_path / "test_data.xlsx"

        generate_excel(config, output_file)

        assert output_file.exists()

        # Verify data was written
        wb = openpyxl.load_workbook(output_file)
        units_sheet = wb['الوحدات']
        assert units_sheet['A2'].value == "101"


class TestCreateSheets:
    """Tests for individual sheet creation functions"""

    def test_create_buildings_sheet(self):
        """Test creating buildings sheet"""
        wb = openpyxl.Workbook()
        del wb['Sheet']  # Remove default sheet

        buildings = [{"name": "عمارة أ", "units": 5, "notes": "test"}]
        create_buildings_sheet(wb, buildings)

        assert 'العمارات' in wb.sheetnames
        sheet = wb['العمارات']
        assert sheet['A1'].value == 'اسم العمارة'
        assert sheet['A2'].value == 'عمارة أ'

    def test_create_units_sheet(self):
        """Test creating units sheet"""
        wb = openpyxl.Workbook()
        del wb['Sheet']

        units = [{
            "unit_no": "101",
            "building": "عمارة أ",
            "type": "شقة",
            "rent": 5000,
            "status": "مُؤجّرة",
            "notes": ""
        }]
        create_units_sheet(wb, units)

        assert 'الوحدات' in wb.sheetnames
        sheet = wb['الوحدات']
        assert sheet['A2'].value == "101"

    def test_create_rents_sheet_with_unpaid(self):
        """Test creating rents sheet with unpaid rent highlighting"""
        wb = openpyxl.Workbook()
        del wb['Sheet']

        rents = [
            {
                "unit_no": "101",
                "month": "يناير",
                "year": 2024,
                "amount": 5000,
                "date": None,
                "method": "",
                "status": "غير مدفوع",
                "notes": ""
            }
        ]
        create_rents_sheet(wb, rents)

        sheet = wb['الإيجارات']
        # Check that unpaid rent row is highlighted (openpyxl uses ARGB format)
        assert sheet['A2'].fill.start_color.rgb == "00FFC7CE"


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
