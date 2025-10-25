#!/usr/bin/env python3
"""
Excel Generator for Building Management
Generates Excel spreadsheets for managing buildings, units, tenants, rents, and expenses.
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def set_cell_style(
    cell: Any,
    font_size: int = 12,
    bold: bool = False,
    bg_color: Optional[str] = None,
    align: str = "center"
) -> None:
    """
    Apply formatting to a cell.

    Args:
        cell: Excel cell to format
        font_size: Font size (default: 12)
        bold: Make font bold (default: False)
        bg_color: Background color in hex format (default: None)
        align: Text alignment - left, center, right (default: center)
    """
    cell.font = Font(size=font_size, bold=bold)
    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal=align, vertical='center')


def load_config(config_path: Path) -> Dict[str, Any]:
    """
    Load configuration from JSON file.

    Args:
        config_path: Path to configuration file

    Returns:
        Dictionary containing configuration data

    Raises:
        FileNotFoundError: If config file doesn't exist
        json.JSONDecodeError: If config file is invalid JSON
    """
    if not config_path.exists():
        raise FileNotFoundError(
            f"Configuration file not found: {config_path}\n"
            f"Please create a config file. See config.example.json for reference."
        )

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(
            f"Invalid JSON in configuration file: {e.msg}",
            e.doc,
            e.pos
        )


def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """
    Parse date string to datetime object.

    Args:
        date_str: Date string in format YYYY-MM-DD

    Returns:
        datetime object or None if date_str is None/empty
    """
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        print(f"Warning: Invalid date format: {date_str}", file=sys.stderr)
        return None


def create_buildings_sheet(wb: Workbook, buildings: List[Dict[str, Any]]) -> None:
    """
    Create the buildings sheet.

    Args:
        wb: Workbook object
        buildings: List of building dictionaries
    """
    sheet = wb.create_sheet('العمارات')
    headers = ['اسم العمارة', 'عدد الوحدات', 'ملاحظات']

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        set_cell_style(cell, bold=True, bg_color="C0C0C0")
        sheet.column_dimensions[get_column_letter(col_num)].width = 15

    # Write data
    for row_num, building in enumerate(buildings, 2):
        sheet.cell(row=row_num, column=1).value = building.get('name', '')
        sheet.cell(row=row_num, column=2).value = building.get('units', 0)
        sheet.cell(row=row_num, column=3).value = building.get('notes', '')

        for col_num in range(1, 4):
            set_cell_style(sheet.cell(row=row_num, column=col_num))


def create_units_sheet(wb: Workbook, units: List[Dict[str, Any]]) -> None:
    """
    Create the units sheet.

    Args:
        wb: Workbook object
        units: List of unit dictionaries
    """
    sheet = wb.create_sheet('الوحدات')
    headers = ['رقم الوحدة', 'العمارة', 'التصنيف', 'الإيجار الشهري', 'الحالة', 'ملاحظات']

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        set_cell_style(cell, bold=True, bg_color="C0C0C0")
        sheet.column_dimensions[get_column_letter(col_num)].width = 15

    # Write data
    for row_num, unit in enumerate(units, 2):
        sheet.cell(row=row_num, column=1).value = unit.get('unit_no', '')
        sheet.cell(row=row_num, column=2).value = unit.get('building', '')
        sheet.cell(row=row_num, column=3).value = unit.get('type', '')
        sheet.cell(row=row_num, column=4).value = unit.get('rent', 0)
        sheet.cell(row=row_num, column=5).value = unit.get('status', '')
        sheet.cell(row=row_num, column=6).value = unit.get('notes', '')

        for col_num in range(1, 7):
            set_cell_style(sheet.cell(row=row_num, column=col_num))


def create_tenants_sheet(wb: Workbook, tenants: List[Dict[str, Any]]) -> None:
    """
    Create the tenants sheet.

    Args:
        wb: Workbook object
        tenants: List of tenant dictionaries
    """
    sheet = wb.create_sheet('المستأجرين')
    headers = [
        'رقم الوحدة', 'اسم المستأجر', 'رقم الهوية', 'رقم الجوال',
        'تاريخ بداية العقد', 'تاريخ نهاية العقد', 'قيمة الإيجار',
        'البريد الإلكتروني', 'ملاحظات'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        set_cell_style(cell, bold=True, bg_color="C0C0C0")
        sheet.column_dimensions[get_column_letter(col_num)].width = 15

    # Write data
    for row_num, tenant in enumerate(tenants, 2):
        sheet.cell(row=row_num, column=1).value = tenant.get('unit_no', '')
        sheet.cell(row=row_num, column=2).value = tenant.get('name', '')
        sheet.cell(row=row_num, column=3).value = tenant.get('id', '')
        sheet.cell(row=row_num, column=4).value = tenant.get('mobile', '')
        sheet.cell(row=row_num, column=5).value = parse_date(tenant.get('start_date'))
        sheet.cell(row=row_num, column=6).value = parse_date(tenant.get('end_date'))
        sheet.cell(row=row_num, column=7).value = tenant.get('rent', 0)
        sheet.cell(row=row_num, column=8).value = tenant.get('email', '')
        sheet.cell(row=row_num, column=9).value = tenant.get('notes', '')

        for col_num in range(1, 10):
            set_cell_style(sheet.cell(row=row_num, column=col_num))


def create_rents_sheet(wb: Workbook, rents: List[Dict[str, Any]]) -> None:
    """
    Create the rents sheet with conditional formatting for unpaid rents.

    Args:
        wb: Workbook object
        rents: List of rent payment dictionaries
    """
    sheet = wb.create_sheet('الإيجارات')
    headers = [
        'رقم الوحدة', 'الشهر', 'السنة', 'قيمة الإيجار',
        'تاريخ الدفع', 'طريقة الدفع', 'الحالة', 'ملاحظات'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        set_cell_style(cell, bold=True, bg_color="C0C0C0")
        sheet.column_dimensions[get_column_letter(col_num)].width = 15

    # Write data
    for row_num, rent in enumerate(rents, 2):
        sheet.cell(row=row_num, column=1).value = rent.get('unit_no', '')
        sheet.cell(row=row_num, column=2).value = rent.get('month', '')
        sheet.cell(row=row_num, column=3).value = rent.get('year', 0)
        sheet.cell(row=row_num, column=4).value = rent.get('amount', 0)
        sheet.cell(row=row_num, column=5).value = parse_date(rent.get('date'))
        sheet.cell(row=row_num, column=6).value = rent.get('method', '')
        sheet.cell(row=row_num, column=7).value = rent.get('status', '')
        sheet.cell(row=row_num, column=8).value = rent.get('notes', '')

        # Apply conditional formatting for unpaid rents
        status = rent.get('status', '')
        bg_color = "FFC7CE" if status == "غير مدفوع" else None

        for col_num in range(1, 9):
            set_cell_style(sheet.cell(row=row_num, column=col_num), bg_color=bg_color)


def create_expenses_sheet(wb: Workbook, expenses: List[Dict[str, Any]]) -> None:
    """
    Create the expenses sheet.

    Args:
        wb: Workbook object
        expenses: List of expense dictionaries
    """
    sheet = wb.create_sheet('المصروفات')
    headers = ['العمارة', 'التاريخ', 'نوع المصروفات', 'القيمة', 'الفئة', 'ملاحظات']

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        set_cell_style(cell, bold=True, bg_color="C0C0C0")
        sheet.column_dimensions[get_column_letter(col_num)].width = 15

    # Write data
    for row_num, expense in enumerate(expenses, 2):
        sheet.cell(row=row_num, column=1).value = expense.get('building', '')
        sheet.cell(row=row_num, column=2).value = parse_date(expense.get('date'))
        sheet.cell(row=row_num, column=3).value = expense.get('type', '')
        sheet.cell(row=row_num, column=4).value = expense.get('amount', 0)
        sheet.cell(row=row_num, column=5).value = expense.get('category', '')
        sheet.cell(row=row_num, column=6).value = expense.get('notes', '')

        for col_num in range(1, 7):
            set_cell_style(sheet.cell(row=row_num, column=col_num))


def generate_excel(config: Dict[str, Any], output_path: Path) -> None:
    """
    Generate Excel file from configuration data.

    Args:
        config: Configuration dictionary with all data
        output_path: Path where Excel file will be saved

    Raises:
        PermissionError: If unable to write to output path
        KeyError: If required keys missing from config
    """
    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create all sheets
    create_buildings_sheet(wb, config.get('buildings', []))
    create_units_sheet(wb, config.get('units', []))
    create_tenants_sheet(wb, config.get('tenants', []))
    create_rents_sheet(wb, config.get('rents_paid', []))
    create_expenses_sheet(wb, config.get('expenses', []))

    # Save workbook
    try:
        wb.save(output_path)
        print(f"✓ Excel file generated successfully: {output_path}")
    except PermissionError:
        raise PermissionError(
            f"Unable to write to {output_path}. "
            f"Please close the file if it's open and try again."
        )


def main() -> int:
    """
    Main entry point for the script.

    Returns:
        Exit code (0 for success, non-zero for errors)
    """
    parser = argparse.ArgumentParser(
        description='Generate Excel spreadsheet for building management',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                           # Use config.json in current directory
  %(prog)s -c myconfig.json          # Use custom config file
  %(prog)s -o output.xlsx            # Specify output filename
  %(prog)s -c data.json -o report.xlsx  # Custom config and output

For configuration format, see config.example.json
        """
    )

    parser.add_argument(
        '-c', '--config',
        type=Path,
        default=Path('config.json'),
        help='Path to configuration JSON file (default: config.json)'
    )

    parser.add_argument(
        '-o', '--output',
        type=Path,
        help='Output Excel file path (default: from config or "output.xlsx")'
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Enable verbose output'
    )

    args = parser.parse_args()

    try:
        # Load configuration
        if args.verbose:
            print(f"Loading configuration from: {args.config}")

        config = load_config(args.config)

        # Determine output path
        if args.output:
            output_path = args.output
        elif 'output_filename' in config:
            output_path = Path(config['output_filename'])
        else:
            output_path = Path('output.xlsx')

        if args.verbose:
            print(f"Output will be saved to: {output_path}")

        # Generate Excel file
        generate_excel(config, output_path)

        return 0

    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON - {e}", file=sys.stderr)
        return 1

    except PermissionError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    except KeyError as e:
        print(f"Error: Missing required key in config: {e}", file=sys.stderr)
        return 1

    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
